import io
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelGeneratorService:
    """Servicio para generar archivos Excel con datos de reportes"""

    def generate_excel_report(self, data: List[Dict[str, Any]], filename: str = "reporte", start_date: str = None, end_date: str = None) -> io.BytesIO:
        """
        Genera un archivo Excel basado en los datos proporcionados.
        
        Args:
            data: Lista de diccionarios con los datos del reporte
            filename: Nombre base del archivo (sin extensión)
            
        Returns:
            io.BytesIO: Stream de bytes del archivo Excel
        """
        if not data:
            # Si no hay datos, crear un Excel vacío con mensaje
            return self._create_empty_excel()
        
        # Crear archivo Excel en memoria
        excel_buffer = io.BytesIO()
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Citas"
        
        # Agregar título
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        
        # Crear título con rango de fechas si están disponibles
        if start_date and end_date:
            try:
                # Convertir fechas a formato DD/MM/YYYY
                start_dt = datetime.strptime(start_date.split(' ')[0], '%Y-%m-%d')
                end_dt = datetime.strptime(end_date.split(' ')[0], '%Y-%m-%d')
                title_cell.value = f"Reporte de Citas del {start_dt.strftime('%d/%m/%Y')} al {end_dt.strftime('%d/%m/%Y')}"
            except (ValueError, AttributeError):
                title_cell.value = f"Reporte de Citas - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        else:
            title_cell.value = f"Reporte de Citas - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Agregar headers en la fila 3
        headers = [
            "ID Cita", "Cliente", "Barbero", "Servicio", "Precio", "Ubicación",
            "Fecha Inicio", "Fecha Fin", "Estado", "Fecha Creación"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Agregar datos
        for row_num, row_data in enumerate(data, 4):
            ws.cell(row=row_num, column=1, value=row_data.get('appointment_id', ''))
            ws.cell(row=row_num, column=2, value=row_data.get('customer_name', ''))
            ws.cell(row=row_num, column=3, value=row_data.get('user_name', ''))
            ws.cell(row=row_num, column=4, value=row_data.get('service_name', ''))
            
            # Celda de precio con formato de moneda peruana
            price_value = row_data.get('service_price', 0)
            # Asegurar que el valor sea numérico
            if isinstance(price_value, (int, float)):
                numeric_price = float(price_value)
            else:
                try:
                    numeric_price = float(price_value) if price_value else 0.0
                except (ValueError, TypeError):
                    numeric_price = 0.0
            
            price_cell = ws.cell(row=row_num, column=5, value=numeric_price)
            price_cell.number_format = '"S/. "#,##0.00'
            
            ws.cell(row=row_num, column=6, value=row_data.get('location_name', ''))
            ws.cell(row=row_num, column=7, value=row_data.get('start_datetime', ''))
            ws.cell(row=row_num, column=8, value=row_data.get('end_datetime', ''))
            ws.cell(row=row_num, column=9, value=row_data.get('status_name', ''))
            ws.cell(row=row_num, column=10, value=row_data.get('insert_date', ''))
        
        # Agregar fila de total si hay datos
        if data:
            total_row = len(data) + 5  # Fila después de todos los datos (comenzamos en fila 4)
            
            # Calcular suma total de precios
            total_amount = sum(
                float(row_data.get('service_price', 0))
                for row_data in data
            )
            
            # Etiqueta "TOTAL"
            total_label_cell = ws.cell(row=total_row, column=4, value="TOTAL:")
            total_label_cell.font = Font(bold=True, size=12)
            total_label_cell.alignment = Alignment(horizontal='right')
            
            # Valor total con formato de moneda
            total_value_cell = ws.cell(row=total_row, column=5, value=total_amount)
            total_value_cell.number_format = '"S/. "#,##0.00'
            total_value_cell.font = Font(bold=True, size=12)
            total_value_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # Agregar tabla de totales por barbero
            barber_totals = {}
            for row_data in data:
                staff_name = row_data.get("user_name", "Sin asignar")
                service_price = row_data.get("service_price", 0.0) or 0.0
                
                try:
                    numeric_price = float(service_price) if service_price else 0.0
                except (ValueError, TypeError):
                    numeric_price = 0.0
                    
                if staff_name in barber_totals:
                    barber_totals[staff_name] += numeric_price
                else:
                    barber_totals[staff_name] = numeric_price

            # Posicionar la tabla de barberos después de un espacio
            barber_table_start_row = total_row + 3
            
            # Encabezados de la tabla de barberos
            header1 = ws.cell(row=barber_table_start_row, column=1, value="Barbero")
            header2 = ws.cell(row=barber_table_start_row, column=2, value="Total")
            
            # Formatear encabezados
            for header_cell in [header1, header2]:
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center')

            # Agregar datos de barberos
            current_barber_row = barber_table_start_row + 1
            for barber_name, barber_total in barber_totals.items():
                ws.cell(row=current_barber_row, column=1, value=barber_name)
                
                total_cell = ws.cell(row=current_barber_row, column=2, value=barber_total)
                total_cell.number_format = '"S/. "#,##0.00'
                
                current_barber_row += 1
        
        # Ajustar ancho de columnas
        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            # Usar el número de columna para obtener la letra, evitando MergedCell
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _create_empty_excel(self) -> io.BytesIO:
        """Crea un archivo Excel vacío con mensaje informativo"""
        excel_buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Vacío"
        # Mensaje de no datos
        ws['A1'] = "No se encontraron datos para los filtros especificados"
        ws['A1'].font = Font(bold=True, size=12)
        
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
